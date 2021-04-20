import os, shutil
from datetime import datetime
import os.path as path
from typing import Type
from pathlib import Path

import fileseq
from openpyxl import load_workbook
from openpyxl.styles import Fill, PatternFill
from openpyxl.utils import get_column_letter

from .base_fittings import IO_Fitting

from ...database.models import (
    Client, Project, Shot, Render, Compo, Grade, WorkingPost
)

class Set_Prod_Number(IO_Fitting):
    def fitting(self):
        self.fso.props.prod_number = (
            int(self.fso.props.season) * 100 + int(self.fso.props.episode)
        )

class Extract_All_Dirs(IO_Fitting):
    def fitting(self):
        self.fso.props.dirs = []
        root_path  = Path(self.fso.path)
        for p in root_path.rglob('**/'):
            if p.is_dir():
                self.fso.props.dirs.append(p)
        

class Generate_Shot_Name(IO_Fitting):
    def fitting(self):
        props = self.fso.props
        self.fso.props.shot_name = (
            f'monster_S{str(int(props.season)).zfill(2)}'
            f'E{str(int(props.episode)).zfill(2)}_'
            f'SQ{str(int(props.sequence)).zfill(4)}_'
            f'SH{str(int(props.shot)).zfill(4)}'
        )



class Detect_IMG_Sequences(IO_Fitting):
    def fitting(self):

        img_seq = None
        for d in self.fso.props.dirs:
            p = d.__str__()
            exr_name = f'{self.fso.props.shot_name}.#.exr'
            exr_ver_name = f'{self.fso.props.shot_name}_v{self.fso.props.version}.#.exr'
            seq_path = path.join(p, exr_name)
            seq_ver_path = path.join(p, exr_ver_name)
            dir_seq = None
            try:
                dir_seq = fileseq.findSequenceOnDisk(seq_path)
            except:
                pass
            try:
                dir_seq = fileseq.findSequenceOnDisk(seq_ver_path)
            except:
                pass
            if not dir_seq:
                break
            img_seq = dir_seq
            # if 
        if img_seq:
            self.fso.props.img_seq = img_seq
        else:
            self.fso.props.error = (
                f'Secuencia de imgágenes no encontrado'
            )
            self.state.raise_error()

class Check_For_Seq_Gaps(IO_Fitting):
    def fitting(self):
        if 'img_seq' in self.fso.props:
            seq = self.fso.props.img_seq
            first_frame = seq.frameSet().start()
            last_frame = seq.frameSet().end()
            expected_num_frames = int(last_frame) - int(first_frame) + 1
            actual_num_frames = len(seq.frameSet())
            if actual_num_frames != expected_num_frames:
                self.fso.props.error = (
                    f'Hay {actual_num_frames} imágenes en la secuencia. Debe ser {expected_num_frames}.')
                self.state.raise_error()



class Get_Project_DB(IO_Fitting):
    def fitting(self):
        client = self.fso.props.client
        prod_num = int(self.fso.props.season) * 100 + int(self.fso.props.episode)
        if not 'project' in self.fso.props:
            self.fso.props.project = Project().new_or_get(
                client = client,
                prod_num = prod_num
            )


class Get_Shot_DB(IO_Fitting):
    def fitting(self):
        if not 'shot_db' in self.fso.props:
            self.fso.props.shot_db = Shot().new_or_get(
                project = self.fso.props.project,
                name = self.fso.props.shot_name,
                shot = int(self.fso.props.shot)
            )


class Get_Version_Number(IO_Fitting):
    def fitting(self):
        if not 'version' in self.fso.props:
            prev_renders = self.fso.props.shot_db.renders
            self.fso.props.version = len(prev_renders) + 1


class Rename_Dir_With_Vers(IO_Fitting):
    def fitting(self):
        if 'version' in self.fso.props:
            new_name = (
                f'{self.fso.props.shot_name}'
                f'_v{str(self.fso.props.version).zfill(3)}'
                # f'{self.fso.extension}'
            )
            new_path = path.join(self.fso.directory, new_name)
            os.rename(self.fso.path, new_path)
            self.fso.path = new_path

class Rename_Seq_With_Vers(IO_Fitting):
    def fitting(self):
        if 'img_seq' in self.fso.props:
            seq = self.fso.props.img_seq
            seq_dir = seq.dirname()
            seq_basename = seq.basename().split(".")[0]
            vers_basename = (
                f'{seq_basename}_v{str(self.fso.props.version).zfill(3)}'
            )



            for idx, frame in enumerate(seq.frameSet()):
                frame_path = seq[idx]
                frame_num = str(frame).zfill(4)
                frame_ext = path.splitext(frame_path)[1]
                frame_vers_basename = (
                    f'{vers_basename}.{frame_num}'
                    f'{frame_ext}'
                )
                frame_vers_path = path.join(seq_dir, frame_vers_basename)

                os.rename(frame_path, frame_vers_path)

            self.fso.props.img_seq.setBasename(f'{vers_basename}.')


class Move_Render_To_Server(IO_Fitting):
    def fitting(self):
        props = self.fso.props
        episode = str(int(props.season) * 100 + int(props.episode)).zfill(3)
        server_dir = path.join(
            props.server,
            props.program,
            str(int(props.season)*100),
            episode,
            'Video',
        )

        render_dir = path.join(server_dir, 'Render')
        if not path.isdir(render_dir):
            os.makedirs(render_dir)
        server_path = path.join(render_dir, self.fso.filename)
        shutil.move(self.fso.path, server_path)
        self.fso.path = server_path
        self.fso.props.server_dir = server_dir


class Generate_Transcode_Dirs(IO_Fitting):
    def fitting(self):
        self.fso.props.qt_dir = path.join(self.fso.props.server_dir, 'Shots')
        self.fso.props.alfa_dir = path.join(self.fso.props.server_dir, 'Shot_Alfas')

        if not path.isdir(self.fso.props.qt_dir):
            os.makedirs(self.fso.props.qt_dir)

        if not path.isdir(self.fso.props.alfa_dir):
            os.makedirs(self.fso.props.alfa_dir)


class Save_Render_To_DB(IO_Fitting):
    def fitting(self):
        self.fso.props.render_db = Render().new_or_get(
            shot = self.fso.props.shot_db,
            location = self.fso.directory,
            name = self.fso.name,
            inbound_name = self.fso.filename,
        )

# class Update_


class Render_Generate_Edit_Dirs(IO_Fitting):
    def fitting(self):
        recursos_dir = path.join(
            self.fso.props.editorial,
            'Recursos',
            'Video',
            str(int(self.fso.props.season)*100),
            str(int(self.fso.props.prod_number))
        )

        self.fso.props.edit_shot_dir = path.join(recursos_dir, 'Shots')
        if not path.isdir(self.fso.props.edit_shot_dir):
            os.makedirs(self.fso.props.edit_shot_dir)

        if 'qt_path' in self.fso.props.qt_path:
            transcode_ext = path.splitext(self.fso.props.qt_path)[1]
        else:
            transcode_ext = '.mov'

        self.fso.props.edit_filename = f'{self.fso.props.shot_name}{transcode_ext}'
    
        self.fso.props.compo_src_dir = path.join(recursos_dir, 'Compo_Src')
        self.fso.props.compo_src_path = path.join(
            self.fso.props.compo_src_dir,
            self.fso.props.edit_filename)

        self.fso.props.edit_shot_path = path.join(
            self.fso.props.edit_shot_dir,
            self.fso.props.edit_filename
        )
        self.fso.path = self.fso.props.qt_path
        self.fso.props.edit_shot_alfa = path.join(recursos_dir, 'Shot_Alfas')
        if not path.isdir(self.fso.props.edit_shot_alfa):
            os.makedirs(self.fso.props.edit_shot_alfa)



class Render_Get_WorkingDB_Shot(IO_Fitting):
    def fitting(self):
        self.fso.props.working_db = WorkingPost().new_or_get(
            shot = self.fso.props.shot_db,
            name = self.fso.props.edit_filename,
            location = self.fso.props.edit_shot_path,
            render = self.fso.props.render_db
        )

class Establish_Working_Dir(IO_Fitting):
    def fitting(self):
        working_db = self.fso.props.working_db
        if (working_db.compo or working_db.grade):
            if working_db.compo and not working_db.grade:
                if not path.isdir(self.fso.props.compo_src_dir):
                    os.makedirs(self.fso.props.compo_src_dir)
                self.fso.props.qt_dir = self.fso.props.compo_src_dir
        else:
            self.fso.props.qt_dir = self.fso.props.edit_shot_dir

# whuh? need to make a transcode path prop
class Update_Working_Shot(IO_Fitting):
    def fitting(self):
        if 'qt_dir' in self.fso.props:
            if not path.isdir(self.fso.props.qt_dir):
                os.makedirs(self.fso.props.qt_dir)
            qt_path = path.join(self.fso.props.qt_dir, self.fso.props.shot_name + '.mov')


            shutil.copy2(self.fso.props.qt_path, qt_path)
            self.fso.props.qt_path = qt_path
            alfa_path = path.join(self.fso.props.edit_shot_alfa, self.fso.props.shot_name + '.mov')
            shutil.copy2(self.fso.props.alfa_path, alfa_path)
        

class Update_WorkingDB(IO_Fitting):
    def fitting(self):
        update = (WorkingPost().update(
            location = self.fso.props.qt_path, 
            modified = datetime.now(),
            render = self.fso.props.render_db    
            ).where(
                WorkingPost().guid == self.fso.props.working_db.guid
            )
        )
        update.execute()

class Update_Excel(IO_Fitting):
    def add_headers(self, sheet):
        sheet['A1'] = 'Plano'
        sheet['B1'] = 'Fecha Ingresada'
        sheet['C1'] = 'Render GUID'
        sheet['D1'] = 'Shot GUID'
        sheet['E1'] = 'Grade'
        sheet['F1'] = 'Compo'
        # sheet['G1'] = 'Alta en editorial'

        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 21
        sheet.column_dimensions['C'].width = 24
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 7
        sheet.column_dimensions['F'].width = 7
        # sheet.column_dimensions['G'].width = 19


    def fitting(self): 
        prod_num = str(self.fso.props.prod_number)
        excel_dir = path.join(self.fso.props.editorial, 'Recursos', 'Docs')
        excel_path = path.join(excel_dir, 'monster-T02-renders.xlsx')
        
        workbook = load_workbook(filename=excel_path)
        
        sheet = None
        sheets = workbook.sheetnames

        if not prod_num in sheets:
            new_sheet_idx = 1
            for _, s in enumerate(sorted(sheets)):
                if int(prod_num) > int(s):
                    new_sheet_idx += 1
                    
            workbook.create_sheet(title=prod_num, index=new_sheet_idx)
            sheet = workbook.get_sheet_by_name(prod_num)
            self.add_headers(sheet)
        else:
            sheet = workbook[prod_num]
        

        row_idx = 2
        other_vers = []     
        for row in sheet.iter_rows(min_row=2):

            if self.fso.name > row[0].value:
                row_idx = row[0].row + 1
            if self.fso.props.shot_db.guid == row[3].value:
                other_vers.append(row)
        
        render_guids = [c.value for c in sheet['C']]
        if not self.fso.props.render_db.guid in render_guids:
            sheet.insert_rows(row_idx, amount=1)
            sheet[f'A{row_idx}'].value = self.fso.name
            sheet[f'B{row_idx}'].value = datetime.now()
            sheet[f'C{row_idx}'].value = self.fso.props.render_db.guid
            sheet[f'D{row_idx}'].value = self.fso.props.shot_db.guid
            if self.fso.props.working_db.grade:
                sheet[f'E{row_idx}'].value = 'sí'
                sheet[f'E{row_idx}'].fill = PatternFill('solid', fgColor='ff85fb')
            if self.fso.props.working_db.compo:
                sheet[f'F{row_idx}'].value = 'sí'
                sheet[f'F{row_idx}'].fill = PatternFill('solid', fgColor='59c8ff')
            # sheet[f'G{row_idx}'].value = 'sí'
            # sheet[f'G{row_idx}'].fill = PatternFill('solid', fgColor='45ff9f')

        try:
            workbook.save(excel_path)
        except:
            temp_filename = (
                f'monster-T02-renders.'
                f'copia{datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.xlsx'
                )
            temp_excel_path = path.join(excel_dir, temp_filename)
            try:
                workbook.save(temp_excel_path)
            except Exception as e:
                print(e)


        


        
        



        


        




        






